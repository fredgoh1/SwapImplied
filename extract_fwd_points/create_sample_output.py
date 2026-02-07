#!/usr/bin/env python3
"""
Demo script - Creates sample Excel output without web scraping
This demonstrates what the actual output will look like
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_excel():
    """Create sample Excel file with demo data"""
    
    # Sample data (based on actual format from Investing.com)
    sample_data = [
        {
            'Period': '1-Month',
            'Bid': '-27.4700',
            'Ask': '-27.1700',
            'High': '-27.4700',
            'Low': '-27.3000',
            'Change': '0.0900',
            'Time': '0:56:29'
        },
        {
            'Period': '3-Month',
            'Bid': '-77.7300',
            'Ask': '-77.3300',
            'High': '-77.6500',
            'Low': '-77.5500',
            'Change': '-0.0100',
            'Time': '0:57:44'
        },
        {
            'Period': '6-Month',
            'Bid': '-148.9700',
            'Ask': '-147.9700',
            'High': '-148.6000',
            'Low': '-148.5000',
            'Change': '0.0300',
            'Time': '0:56:14'
        }
    ]
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Forward Points'
    
    # Title
    sheet['A1'] = 'USD/SGD Forward Points'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet.merge_cells('A1:G1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    sheet['A2'] = f'Sample Data - Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    sheet['A2'].font = Font(size=10, italic=True, color='666666')
    sheet.merge_cells('A2:G2')
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Period', 'Bid', 'Ask', 'High', 'Low', 'Change', 'Time']
    header_row = 4
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        cell.border = thin_border
    
    # Data rows
    for row_num, record in enumerate(sample_data, header_row + 1):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = record[header]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Border
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            cell.border = thin_border
            
            # Bold period names
            if col_num == 1:
                cell.font = Font(bold=True, size=10)
    
    # Column widths
    column_widths = {
        'A': 15,
        'B': 12,
        'C': 12,
        'D': 12,
        'E': 12,
        'F': 12,
        'G': 18
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Notes
    note_row = header_row + len(sample_data) + 2
    sheet[f'A{note_row}'] = 'Note: Forward points are in pips. Negative values indicate forward discount.'
    sheet[f'A{note_row}'].font = Font(size=9, italic=True, color='666666')
    
    sheet[f'A{note_row + 1}'] = 'Source: Investing.com (https://www.investing.com/currencies/usd-sgd-forward-rates)'
    sheet[f'A{note_row + 1}'].font = Font(size=9, italic=True, color='666666')
    
    sheet[f'A{note_row + 2}'] = 'This is SAMPLE DATA for demonstration purposes only.'
    sheet[f'A{note_row + 2}'].font = Font(size=9, bold=True, italic=True, color='FF0000')
    
    # Freeze panes
    sheet.freeze_panes = f'A{header_row + 1}'
    
    filename = 'SAMPLE_usd_sgd_forward_points.xlsx'
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    print("\nThis demonstrates the format of the actual output.")
    print("The actual script will fetch live data from Investing.com")

if __name__ == "__main__":
    create_sample_excel()
