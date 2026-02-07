#!/usr/bin/env python3
"""
Extract USD/SGD Forward Points from Investing.com or Browse.AI

Supports three modes:
  1. HTTP requests (default) - scrapes investing.com directly
  2. Selenium (--selenium) - uses headless browser for more reliable scraping
  3. Browse.AI (--browse-ai) - triggers Browse.AI robot, parses capturedLists,
     then updates master files and calculates swap implied rates end-to-end

Usage:
    python extract_forward_points.py                # requests mode
    python extract_forward_points.py --selenium     # selenium mode
    python extract_forward_points.py --browse-ai    # full pipeline via Browse.AI
"""

import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os
import argparse
import pandas as pd
from pathlib import Path

# Add parent directory to path for cross-module imports
_project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_project_root / 'extract_fwd_points'))
sys.path.insert(0, str(_project_root / 'extract_all_rates'))
sys.path.insert(0, str(_project_root / 'calc_swap_implied'))

from browse_ai_extractor import BrowseAIClient, load_credentials
from update_swap_implied_data import DataExtractor, DataUpdater
from calculate_swap_implied_rates import HolidayCalendar, SwapImpliedRateCalculator, find_sofr_column


def extract_forward_points():
    """Extract forward points from Investing.com"""
    url = "https://www.investing.com/currencies/usd-sgd-forward-rates"

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the forward rates table
    table = soup.find('table')
    if not table:
        print("Error: Forward rates table not found")
        return None

    # Extract data for 1M, 3M, and 6M
    target_periods = {
        'USDSGD 1M FWD': '1-Month',
        'USDSGD 3M FWD': '3-Month',
        'USDSGD 6M FWD': '6-Month'
    }

    results = []
    rows = table.find_all('tr')

    for row in rows:
        cells = row.find_all('td')
        if len(cells) >= 7:
            name_cell = cells[0].get_text(strip=True)

            for target_name, display_name in target_periods.items():
                if target_name in name_cell:
                    try:
                        bid = cells[1].get_text(strip=True)
                        ask = cells[2].get_text(strip=True)
                        high = cells[3].get_text(strip=True)
                        low = cells[4].get_text(strip=True)
                        change = cells[5].get_text(strip=True)
                        time = cells[6].get_text(strip=True)

                        results.append({
                            'Period': display_name,
                            'Bid': bid,
                            'Ask': ask,
                            'High': high,
                            'Low': low,
                            'Change': change,
                            'Time': time
                        })
                    except Exception as e:
                        print(f"Error parsing {display_name}: {e}")

    return results

def create_excel(data, filename='usd_sgd_forward_points.xlsx'):
    """Create formatted Excel file with forward points data"""

    if not data:
        print("No data to save")
        return False

    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Forward Points'

    # Title
    sheet['A1'] = 'USD/SGD Forward Points'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet['A2'] = f'Extracted: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    sheet['A2'].font = Font(size=10, italic=True)

    # Headers
    headers = ['Period', 'Bid', 'Ask', 'High', 'Low', 'Change', 'Time']
    header_row = 4

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border

    # Data rows
    for row_num, record in enumerate(data, header_row + 1):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = record[header]
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

            # Highlight period names
            if col_num == 1:
                cell.font = Font(bold=True)

    # Column widths
    column_widths = {
        'A': 15,
        'B': 12,
        'C': 12,
        'D': 12,
        'E': 12,
        'F': 12,
        'G': 15
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Note
    note_row = header_row + len(data) + 2
    sheet[f'A{note_row}'] = 'Note: Forward points are in pips. Negative values indicate forward discount.'
    sheet[f'A{note_row}'].font = Font(size=9, italic=True)

    sheet[f'A{note_row + 1}'] = 'Source: Investing.com (https://www.investing.com/currencies/usd-sgd-forward-rates)'
    sheet[f'A{note_row + 1}'].font = Font(size=9, italic=True)

    try:
        wb.save(filename)
        print(f"Data successfully saved to {filename}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def extract_with_browse_ai(credentials_path, max_wait=300, poll_interval=5):
    """
    Extract forward points via Browse.AI robot.

    Triggers the robot, waits for completion, then parses capturedLists
    to extract bid/ask for each tenor and calculate mid points.

    Args:
        credentials_path: Path to Browse_AI credentials file
        max_wait: Maximum seconds to wait for task completion
        poll_interval: Seconds between status checks

    Returns:
        dict: {'1M': mid_1m, '3M': mid_3m, '6M': mid_6m} or None on failure
    """
    print("=" * 70)
    print("BROWSE.AI FORWARD POINTS EXTRACTION")
    print("=" * 70)
    print()

    # Load credentials
    creds_path = Path(credentials_path)
    if not creds_path.is_absolute():
        creds_path = Path(__file__).parent / creds_path

    print(f"Loading credentials from: {creds_path}")
    api_key, workspace_id, robot_id = load_credentials(creds_path)
    print(f"  Robot ID: {robot_id}")
    print()

    # Initialize client and run task
    client = BrowseAIClient(api_key, robot_id)

    print("-" * 70)
    task_result = client.run_task()
    task_id = task_result.get("id")

    print("-" * 70)
    task = client.wait_for_completion(
        task_id,
        poll_interval=poll_interval,
        max_wait=max_wait
    )

    # Parse capturedLists for forward points
    print()
    print("-" * 70)
    print("PARSING FORWARD POINTS FROM CAPTURED DATA")
    print("-" * 70)

    captured_lists = task.get("capturedLists", {})
    if not captured_lists:
        print("No capturedLists found in task result.")
        # Fall back to capturedTexts if available
        captured_texts = task.get("capturedTexts", {})
        if captured_texts:
            print("Found capturedTexts, but capturedLists parsing is required.")
            print(f"  capturedTexts: {captured_texts}")
        return None

    # Target tenor names in the captured data
    tenor_names = {
        'USDSGD 1M FWD': '1M',
        'USDSGD 3M FWD': '3M',
        'USDSGD 6M FWD': '6M',
    }

    forward_points = {}

    # capturedLists is typically a dict of list_name -> list of row dicts
    for list_name, rows in captured_lists.items():
        print(f"\n  Processing list: {list_name}")
        if not isinstance(rows, list):
            print(f"    Skipping (not a list): {type(rows)}")
            continue

        for row in rows:
            # Row is a dict with column names as keys
            row_name = ""
            # Try common column names for the instrument name
            for name_key in ["Name", "name", "Instrument", "instrument", "Contract", "contract"]:
                if name_key in row:
                    row_name = str(row[name_key]).strip()
                    break

            # If no named column, check first value
            if not row_name and isinstance(row, dict):
                first_val = next(iter(row.values()), "")
                row_name = str(first_val).strip()

            for target_name, tenor in tenor_names.items():
                if target_name in row_name:
                    # Extract Bid and Ask
                    bid = None
                    ask = None
                    for key, val in row.items():
                        key_lower = key.lower()
                        if 'bid' in key_lower:
                            try:
                                bid = float(str(val).replace(',', ''))
                            except (ValueError, TypeError):
                                pass
                        elif 'ask' in key_lower:
                            try:
                                ask = float(str(val).replace(',', ''))
                            except (ValueError, TypeError):
                                pass

                    if bid is not None and ask is not None:
                        mid = round((bid + ask) / 2, 4)
                        forward_points[tenor] = mid
                        print(f"    {tenor}: Bid={bid}, Ask={ask}, Mid={mid}")
                    else:
                        print(f"    {tenor}: Could not parse Bid/Ask (bid={bid}, ask={ask})")
                        print(f"    Row data: {row}")

    if len(forward_points) == 3:
        print(f"\nForward points extracted: 1M={forward_points['1M']}, 3M={forward_points['3M']}, 6M={forward_points['6M']}")
        return forward_points
    else:
        print(f"\nCould not extract all forward points. Found: {list(forward_points.keys())}")
        return None


def update_master_files(forward_points, input_dir=None):
    """
    Extract SOFR rates and FX spot, then append today's data to master files.

    Args:
        forward_points: dict with keys '1M', '3M', '6M'
        input_dir: path to directory containing input_master_*.xlsx files

    Returns:
        tuple: (sofr_rates, fx_rate, update_results) or (None, None, None) on failure
    """
    if input_dir is None:
        input_dir = str(Path(__file__).parent.parent / 'swap_implied_input')

    print()
    print("=" * 70)
    print("UPDATING MASTER FILES")
    print("=" * 70)
    print()

    # Extract SOFR rates and FX spot using existing DataExtractor
    extractor = DataExtractor(use_selenium=False)
    sofr_rates = extractor.extract_sofr_rates()
    fx_rate = extractor.extract_usdsgd_fx()

    if not sofr_rates or not fx_rate:
        print("Failed to extract SOFR rates or FX rate")
        return None, None, None

    # Update master files
    updater = DataUpdater(input_dir=input_dir)
    if not updater.validate_files():
        print(f"Master files not found in {input_dir}")
        return None, None, None

    results = updater.update_files(sofr_rates, fx_rate, forward_points)

    # Summary
    print()
    success_count = sum(1 for r in results.values() if r == 'success')
    print(f"Successfully updated: {success_count}/3 files")
    for period, status in results.items():
        symbol = "+" if status == "success" else "~" if status == "skipped" else "x"
        print(f"  [{symbol}] {period}: {status}")

    return sofr_rates, fx_rate, results


def calculate_implied_rates(input_dir=None, output_dir=None):
    """
    Calculate swap implied SGD rates for all tenors from master files.

    For each tenor (1M, 3M, 6M):
      - Reads input_master_{tenor}.xlsx
      - Calculates implied SGD rate for every row
      - Outputs output_master_{tenor}.xlsx with Date and Implied_SGD_Rate_Pct

    Args:
        input_dir: path to directory containing input_master_*.xlsx files
        output_dir: path to directory for output files (defaults to project root)
    """
    if input_dir is None:
        input_dir = str(Path(__file__).parent.parent / 'swap_implied_input')
    if output_dir is None:
        output_dir = str(Path(__file__).parent.parent)

    print()
    print("=" * 70)
    print("CALCULATING SWAP IMPLIED RATES")
    print("=" * 70)
    print()

    input_path = Path(input_dir)
    output_path = Path(output_dir)

    for tenor in ['1M', '3M', '6M']:
        input_file = input_path / f'input_master_{tenor.lower()}.xlsx'
        output_file = output_path / f'output_master_{tenor.lower()}.xlsx'

        if not input_file.exists():
            print(f"Skipping {tenor}: {input_file} not found")
            continue

        print(f"\nProcessing {tenor} tenor...")
        print(f"  Input:  {input_file}")
        print(f"  Output: {output_file}")

        df = pd.read_excel(input_file)
        if df.empty:
            print(f"  No data in {input_file}")
            continue

        # Find SOFR column
        sofr_col = find_sofr_column(df, tenor)
        if sofr_col is None:
            print(f"  Could not find SOFR column for {tenor}. Columns: {df.columns.tolist()}")
            continue

        # Find FX and forward points columns
        fx_col = None
        fwd_pts_col = None
        for col in df.columns:
            col_upper = col.upper()
            if 'USDSGD' in col_upper or 'FX' in col_upper:
                fx_col = col
            if 'FORWARD' in col_upper or 'FORWARDPOINTS' in col_upper.replace(' ', ''):
                fwd_pts_col = col
        if fx_col is None or fwd_pts_col is None:
            print(f"  Could not find FX or ForwardPoints column. Columns: {df.columns.tolist()}")
            continue

        # Initialize calendar and calculator
        first_year = pd.to_datetime(df['Date'].iloc[0]).year
        calendar = HolidayCalendar(year=first_year)
        calculator = SwapImpliedRateCalculator(calendar, tenor=tenor)

        # Calculate implied rate for each row
        dates = []
        implied_rates = []
        for _, row in df.iterrows():
            trade_date = pd.to_datetime(row['Date'])
            try:
                result = calculator.process_row(
                    trade_date=trade_date,
                    sofr_rate=row[sofr_col],
                    spot_rate=row[fx_col],
                    forward_points=row[fwd_pts_col],
                    verbose=False
                )
                dates.append(trade_date.strftime('%Y-%m-%d'))
                implied_rates.append(result['Implied_SGD_Rate_Pct'])
            except Exception as e:
                print(f"  Error processing {trade_date}: {e}")
                dates.append(trade_date.strftime('%Y-%m-%d'))
                implied_rates.append(None)

        # Write output
        output_df = pd.DataFrame({
            'Date': dates,
            'Implied_SGD_Rate_Pct': implied_rates
        })
        output_df.to_excel(output_file, index=False)
        print(f"  Wrote {len(output_df)} rows to {output_file}")

    print()
    print("=" * 70)
    print("CALCULATION COMPLETE")
    print("=" * 70)


def main():
    parser = argparse.ArgumentParser(
        description='Extract USD/SGD forward points and optionally run full pipeline',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract_forward_points.py                # requests mode (investing.com)
  python extract_forward_points.py --selenium     # selenium mode
  python extract_forward_points.py --browse-ai    # full pipeline via Browse.AI
        """
    )
    parser.add_argument(
        '--selenium',
        action='store_true',
        help='Use Selenium for web scraping (more reliable but slower)'
    )
    parser.add_argument(
        '--browse-ai',
        action='store_true',
        help='Use Browse.AI to extract forward points, then update master files and calculate implied rates'
    )
    parser.add_argument(
        '--update-master',
        action='store_true',
        help='Update master files with extracted data (auto-enabled with --browse-ai)'
    )
    parser.add_argument(
        '--calc-implied',
        action='store_true',
        help='Calculate implied SGD rates from master files (auto-enabled with --browse-ai)'
    )
    parser.add_argument(
        '--credentials',
        default='../Browse_AI',
        help='Path to Browse.AI credentials file (default: ../Browse_AI)'
    )
    parser.add_argument(
        '--input-dir',
        default=None,
        help='Directory containing master Excel files (default: ../swap_implied_input)'
    )
    parser.add_argument(
        '--output-dir',
        default=None,
        help='Directory for output files (default: project root)'
    )
    parser.add_argument(
        '--max-wait',
        type=int,
        default=300,
        help='Maximum seconds to wait for Browse.AI task (default: 300)'
    )
    parser.add_argument(
        '--poll-interval',
        type=int,
        default=5,
        help='Seconds between Browse.AI status checks (default: 5)'
    )

    args = parser.parse_args()

    # --browse-ai auto-enables --update-master and --calc-implied
    if args.browse_ai:
        args.update_master = True
        args.calc_implied = True

    if args.browse_ai:
        # Full pipeline: Browse.AI -> update master -> calculate implied
        forward_points = extract_with_browse_ai(
            credentials_path=args.credentials,
            max_wait=args.max_wait,
            poll_interval=args.poll_interval
        )

        if not forward_points:
            print("\nFailed to extract forward points from Browse.AI")
            return 1

        # Update master files
        sofr_rates, fx_rate, update_results = update_master_files(
            forward_points, input_dir=args.input_dir
        )

        if update_results is None:
            print("\nFailed to update master files")
            return 1

        # Calculate implied rates
        calculate_implied_rates(input_dir=args.input_dir, output_dir=args.output_dir)

        return 0

    elif args.selenium:
        # Selenium mode - use DataExtractor from update_swap_implied_data
        extractor = DataExtractor(use_selenium=True)
        fp = extractor.extract_forward_points()

        if fp:
            print(f"\nExtracted forward points (mid): 1M={fp['1M']}, 3M={fp['3M']}, 6M={fp['6M']}")

            if args.update_master:
                update_master_files(fp, input_dir=args.input_dir)
            if args.calc_implied:
                calculate_implied_rates(input_dir=args.input_dir, output_dir=args.output_dir)
            return 0
        else:
            print("\nFailed to extract forward points")
            return 1

    else:
        # Default: requests mode (original behavior)
        print("Extracting USD/SGD forward points from Investing.com...")

        data = extract_forward_points()

        if data:
            print(f"\nExtracted {len(data)} forward rate entries:")
            for entry in data:
                print(f"  {entry['Period']}: Bid={entry['Bid']}, Ask={entry['Ask']}")

            output_file = 'usd_sgd_forward_points.xlsx'
            if create_excel(data, output_file):
                print(f"\nExcel file created: {output_file}")

                # If --update-master or --calc-implied, compute mid points and proceed
                if args.update_master or args.calc_implied:
                    fp = {}
                    tenor_map = {'1-Month': '1M', '3-Month': '3M', '6-Month': '6M'}
                    for entry in data:
                        tenor = tenor_map.get(entry['Period'])
                        if tenor:
                            try:
                                bid = float(entry['Bid'])
                                ask = float(entry['Ask'])
                                fp[tenor] = round((bid + ask) / 2, 4)
                            except (ValueError, TypeError):
                                pass

                    if len(fp) == 3:
                        if args.update_master:
                            update_master_files(fp, input_dir=args.input_dir)
                        if args.calc_implied:
                            calculate_implied_rates(input_dir=args.input_dir, output_dir=args.output_dir)
                    else:
                        print("Could not compute mid points from extracted data")

                return 0
            else:
                print("\nFailed to create Excel file")
                return 1
        else:
            print("\nFailed to extract forward points data")
            return 1

if __name__ == "__main__":
    sys.exit(main())
