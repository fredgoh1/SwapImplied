#!/usr/bin/env python3
"""
Extract USD/SGD Forward Points from Investing.com
Supports both requests and Selenium for web scraping.
Extracts 1-month, 3-month, and 6-month forward points and saves to Excel.

Usage:
    python extract_forward_points_selenium.py              # Uses requests (faster)
    python extract_forward_points_selenium.py --selenium   # Uses Selenium (more reliable)

Requirements:
    pip install requests beautifulsoup4 openpyxl selenium
    
    For Selenium mode, also install:
    - Chrome browser
    - ChromeDriver (or use webdriver-manager: pip install webdriver-manager)
"""

import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import time

def extract_with_requests():
    """Extract using requests + BeautifulSoup (faster, may be blocked)"""
    import requests
    from bs4 import BeautifulSoup
    
    url = "https://www.investing.com/currencies/usd-sgd-forward-rates"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    }
    
    try:
        print("Fetching data with requests...")
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        return parse_table(soup)
        
    except Exception as e:
        print(f"Error with requests method: {e}")
        return None

def extract_with_selenium():
    """Extract using Selenium (more reliable, slower)"""
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from bs4 import BeautifulSoup
    
    try:
        # Try to use webdriver-manager for automatic ChromeDriver installation
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.chrome.service import Service
            use_manager = True
        except ImportError:
            use_manager = False
            print("webdriver-manager not found. Using system ChromeDriver.")
    except Exception:
        use_manager = False
    
    url = "https://www.investing.com/currencies/usd-sgd-forward-rates"
    
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    driver = None
    try:
        print("Launching Chrome browser...")
        if use_manager:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        else:
            driver = webdriver.Chrome(options=chrome_options)
        
        print(f"Loading {url}...")
        driver.get(url)
        
        # Wait for table to load
        wait = WebDriverWait(driver, 15)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        
        # Give extra time for dynamic content
        time.sleep(2)
        
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        
        return parse_table(soup)
        
    except Exception as e:
        print(f"Error with Selenium method: {e}")
        return None
    finally:
        if driver:
            driver.quit()

def parse_table(soup):
    """Parse the forward rates table from BeautifulSoup object"""
    table = soup.find('table')
    if not table:
        print("Error: Forward rates table not found in page")
        return None
    
    # Target periods we want to extract
    target_periods = {
        'USDSGD 1M FWD': '1-Month',
        'USDSGD 3M FWD': '3-Month',
        'USDSGD 6M FWD': '6-Month'
    }
    
    results = []
    rows = table.find_all('tr')
    
    for row in rows:
        cells = row.find_all('td')
        if len(cells) >= 7:
            name_cell = cells[0].get_text(strip=True)
            
            for target_name, display_name in target_periods.items():
                if target_name in name_cell:
                    try:
                        bid = cells[1].get_text(strip=True)
                        ask = cells[2].get_text(strip=True)
                        high = cells[3].get_text(strip=True)
                        low = cells[4].get_text(strip=True)
                        change = cells[5].get_text(strip=True)
                        time_str = cells[6].get_text(strip=True)
                        
                        results.append({
                            'Period': display_name,
                            'Bid': bid,
                            'Ask': ask,
                            'High': high,
                            'Low': low,
                            'Change': change,
                            'Time': time_str
                        })
                        print(f"✓ Found {display_name}: Bid={bid}, Ask={ask}")
                    except Exception as e:
                        print(f"Error parsing {display_name}: {e}")
    
    if len(results) == 0:
        print("Warning: No matching forward rate entries found")
    
    return results

def create_excel(data, filename='usd_sgd_forward_points.xlsx'):
    """Create formatted Excel file with forward points data"""
    
    if not data or len(data) == 0:
        print("No data to save")
        return False
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Forward Points'
    
    # Title
    sheet['A1'] = 'USD/SGD Forward Points'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet.merge_cells('A1:G1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    sheet['A2'] = f'Extracted: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    sheet['A2'].font = Font(size=10, italic=True, color='666666')
    sheet.merge_cells('A2:G2')
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Period', 'Bid', 'Ask', 'High', 'Low', 'Change', 'Time']
    header_row = 4
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        cell.border = thin_border
    
    # Data rows
    for row_num, record in enumerate(data, header_row + 1):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = record[header]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Border
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            cell.border = thin_border
            
            # Bold period names
            if col_num == 1:
                cell.font = Font(bold=True, size=10)
    
    # Column widths
    column_widths = {
        'A': 15,
        'B': 12,
        'C': 12,
        'D': 12,
        'E': 12,
        'F': 12,
        'G': 18
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Notes
    note_row = header_row + len(data) + 2
    sheet[f'A{note_row}'] = 'Note: Forward points are in pips. Negative values indicate forward discount.'
    sheet[f'A{note_row}'].font = Font(size=9, italic=True, color='666666')
    
    sheet[f'A{note_row + 1}'] = 'Source: Investing.com (https://www.investing.com/currencies/usd-sgd-forward-rates)'
    sheet[f'A{note_row + 1}'].font = Font(size=9, italic=True, color='666666')
    
    # Freeze panes (freeze header row)
    sheet.freeze_panes = f'A{header_row + 1}'
    
    try:
        wb.save(filename)
        print(f"\n✓ Data successfully saved to {filename}")
        return True
    except Exception as e:
        print(f"\n✗ Error saving Excel file: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(
        description='Extract USD/SGD forward points from Investing.com',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--selenium',
        action='store_true',
        help='Use Selenium instead of requests (more reliable but slower)'
    )
    parser.add_argument(
        '-o', '--output',
        default='usd_sgd_forward_points.xlsx',
        help='Output Excel filename (default: usd_sgd_forward_points.xlsx)'
    )
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("USD/SGD Forward Points Extractor")
    print("=" * 60)
    
    # Extract data
    if args.selenium:
        print("\nUsing Selenium method...")
        data = extract_with_selenium()
    else:
        print("\nUsing requests method...")
        data = extract_with_requests()
        
        # Fallback to Selenium if requests fails
        if data is None:
            print("\nRequests method failed. Trying Selenium...")
            data = extract_with_selenium()
    
    # Create Excel file
    if data and len(data) > 0:
        print(f"\n✓ Successfully extracted {len(data)} forward rate entries")
        
        if create_excel(data, args.output):
            print(f"\n{'=' * 60}")
            print(f"SUCCESS! Excel file created: {args.output}")
            print(f"{'=' * 60}")
            return 0
        else:
            return 1
    else:
        print("\n✗ Failed to extract forward points data")
        print("\nTroubleshooting tips:")
        print("1. Check your internet connection")
        print("2. Try running with --selenium flag")
        print("3. Verify the Investing.com website is accessible")
        print("4. Install Chrome and ChromeDriver for Selenium mode")
        return 1

if __name__ == "__main__":
    sys.exit(main())
