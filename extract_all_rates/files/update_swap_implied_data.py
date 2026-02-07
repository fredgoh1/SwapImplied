#!/usr/bin/env python3
"""
Swap-Implied Rate Data Updater
===============================
Extracts SOFR rates, USD/SGD FX rates, and forward points, then appends to master files.

Data Sources:
- SOFR Rates: https://www.global-rates.com/en/interest-rates/cme-term-sofr/
- Forward Points: https://www.investing.com/currencies/usd-sgd-forward-rates
- FX Rates: Multiple sources (alphavantage, exchangerate-api, etc.)

Usage:
    python update_swap_implied_data.py [--selenium] [--input-dir PATH]
"""

import argparse
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from pathlib import Path
import sys
import time

class DataExtractor:
    """Handles extraction of financial data from various sources"""
    
    def __init__(self, use_selenium=False):
        self.use_selenium = use_selenium
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
        }
    
    def extract_sofr_rates(self):
        """
        Extract 1M, 3M, and 6M SOFR rates from global-rates.com
        Returns: dict with keys '1M', '3M', '6M'
        """
        url = "https://www.global-rates.com/en/interest-rates/cme-term-sofr/"
        
        try:
            print("Extracting SOFR rates...")
            
            if self.use_selenium:
                print("  Using Selenium mode (may take 15-30 seconds)...")
                soup = self._fetch_with_selenium(url)
            else:
                print("  Using requests mode (fast)...")
                try:
                    response = requests.get(url, headers=self.headers, timeout=15)
                    response.raise_for_status()
                    soup = BeautifulSoup(response.content, 'html.parser')
                except requests.exceptions.RequestException as e:
                    print(f"  Request failed: {e}")
                    print(f"  Tip: Try running with --selenium flag")
                    raise
            
            if not soup:
                raise Exception("Failed to fetch page")
            
            # Find the SOFR rates table
            sofr_rates = {}
            table = soup.find('table', {'class': 'tablesorter'})
            
            if not table:
                # Try alternative table structures
                print("  Looking for SOFR table...")
                tables = soup.find_all('table')
                for t in tables:
                    if 'SOFR' in t.get_text() or 'sofr' in t.get_text():
                        table = t
                        break
            
            if not table:
                raise Exception("SOFR rates table not found. Website structure may have changed.")
            
            if table:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 2:
                        tenor = cells[0].get_text(strip=True).lower()
                        rate = cells[1].get_text(strip=True)
                        
                        # Match exact tenors - be specific to avoid matching "1" in "12"
                        # Look for "1 month" but NOT "12 month"
                        if ('1 month' in tenor or '1-month' in tenor or '1month' in tenor) and '12' not in tenor:
                            sofr_rates['1M'] = self._clean_rate(rate)
                            print(f"  ✓ 1M: {sofr_rates['1M']}%")
                        elif '3 month' in tenor or '3-month' in tenor or '3month' in tenor:
                            sofr_rates['3M'] = self._clean_rate(rate)
                            print(f"  ✓ 3M: {sofr_rates['3M']}%")
                        elif '6 month' in tenor or '6-month' in tenor or '6month' in tenor:
                            sofr_rates['6M'] = self._clean_rate(rate)
                            print(f"  ✓ 6M: {sofr_rates['6M']}%")
            
            if len(sofr_rates) == 3:
                print(f"✓ SOFR rates extracted: 1M={sofr_rates['1M']}%, 3M={sofr_rates['3M']}%, 6M={sofr_rates['6M']}%")
                return sofr_rates
            else:
                raise Exception(f"Could not extract all SOFR rates. Found: {list(sofr_rates.keys())}")
                
        except Exception as e:
            print(f"✗ Error extracting SOFR rates: {e}")
            if not self.use_selenium and "timeout" not in str(e).lower():
                print("  Tip: Try running with --selenium flag")
            return None
    
    def extract_forward_points(self):
        """
        Extract 1M, 3M, and 6M USD/SGD forward points from investing.com
        Returns mid rates (average of bid and ask)
        Returns: dict with keys '1M', '3M', '6M'
        """
        url = "https://www.investing.com/currencies/usd-sgd-forward-rates"
        
        try:
            print("Extracting forward points...")
            
            if self.use_selenium:
                print("  Using Selenium mode (may take 15-30 seconds)...")
                soup = self._fetch_with_selenium(url, wait_for_table=True)
            else:
                print("  Using requests mode (fast)...")
                try:
                    response = requests.get(url, headers=self.headers, timeout=15)
                    response.raise_for_status()
                    soup = BeautifulSoup(response.content, 'html.parser')
                except requests.exceptions.RequestException as e:
                    print(f"  Request failed: {e}")
                    print(f"  Tip: Try running with --selenium flag for more reliable extraction")
                    raise
            
            if not soup:
                raise Exception("Failed to fetch page")
            
            table = soup.find('table')
            if not table:
                raise Exception("Forward rates table not found. Website structure may have changed.")
            
            target_periods = {
                'USDSGD 1M FWD': '1M',
                'USDSGD 3M FWD': '3M',
                'USDSGD 6M FWD': '6M'
            }
            
            forward_points = {}
            rows = table.find_all('tr')
            
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 3:
                    name = cells[0].get_text(strip=True)
                    
                    for target_name, period in target_periods.items():
                        if target_name in name:
                            try:
                                bid_text = cells[1].get_text(strip=True)
                                ask_text = cells[2].get_text(strip=True)
                                
                                bid = float(bid_text)
                                ask = float(ask_text)
                                mid = (bid + ask) / 2
                                forward_points[period] = round(mid, 4)
                                print(f"  ✓ {period}: Bid={bid}, Ask={ask}, Mid={mid}")
                            except (ValueError, IndexError) as e:
                                print(f"  Warning: Could not parse {period} forward points: {e}")
            
            if len(forward_points) == 3:
                print(f"✓ Forward points extracted: 1M={forward_points['1M']}, 3M={forward_points['3M']}, 6M={forward_points['6M']}")
                return forward_points
            else:
                raise Exception(f"Could not extract all forward points. Found: {list(forward_points.keys())}")
                
        except Exception as e:
            print(f"✗ Error extracting forward points: {e}")
            if not self.use_selenium and "timeout" not in str(e).lower():
                print("  Tip: Try running with --selenium flag")
            return None
    
    def extract_usdsgd_fx(self):
        """
        Extract current USD/SGD FX rate from multiple sources
        Returns: float
        """
        print("Extracting USD/SGD FX rate...")
        
        # Try multiple sources in order of preference
        sources = [
            self._get_fx_from_exchangerate_api,
            self._get_fx_from_fixer,
            self._get_fx_from_xe,
        ]
        
        for source_func in sources:
            try:
                fx_rate = source_func()
                if fx_rate:
                    print(f"✓ USD/SGD FX rate: {fx_rate}")
                    return fx_rate
            except Exception as e:
                print(f"  Trying next source... ({e})")
                continue
        
        print("✗ Could not extract FX rate from any source")
        return None
    
    def _get_fx_from_exchangerate_api(self):
        """Get FX rate from exchangerate-api.com (free, no key required)"""
        url = "https://open.er-api.com/v6/latest/USD"
        response = requests.get(url, timeout=10)
        data = response.json()
        if data.get('result') == 'success' and 'SGD' in data.get('rates', {}):
            return round(data['rates']['SGD'], 4)
        return None
    
    def _get_fx_from_fixer(self):
        """Get FX rate from fixer.io (requires free API key)"""
        # Note: User would need to set their API key
        api_key = None  # User should set this
        if not api_key:
            return None
        
        url = f"http://data.fixer.io/api/latest?access_key={api_key}&symbols=SGD"
        response = requests.get(url, timeout=10)
        data = response.json()
        if data.get('success') and 'SGD' in data.get('rates', {}):
            # Fixer uses EUR as base, need to convert to USD
            eur_sgd = data['rates']['SGD']
            eur_usd = data['rates'].get('USD', 1.0)
            return round(eur_sgd / eur_usd, 4)
        return None
    
    def _get_fx_from_xe(self):
        """Scrape FX rate from XE.com"""
        url = "https://www.xe.com/currencyconverter/convert/?Amount=1&From=USD&To=SGD"
        response = requests.get(url, headers=self.headers, timeout=15)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Look for the conversion result
        result = soup.find('p', {'class': 'result__BigRate-sc-1bsijpp-1'})
        if result:
            rate_text = result.get_text(strip=True)
            # Extract number from text like "1.35 SGD"
            rate = float(''.join(c for c in rate_text.split()[0] if c.isdigit() or c == '.'))
            return round(rate, 4)
        return None
    
    def _fetch_with_selenium(self, url, wait_for_table=False):
        """Fetch page using Selenium"""
        try:
            from selenium import webdriver
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.chrome.options import Options
            from selenium.common.exceptions import TimeoutException, WebDriverException
        except ImportError as e:
            raise Exception(f"Selenium not installed. Install with: pip install selenium webdriver-manager")
        
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.chrome.service import Service
            use_manager = True
        except ImportError:
            use_manager = False
        
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-logging')
        chrome_options.add_argument('--log-level=3')
        chrome_options.add_argument(f'--user-agent={self.headers["User-Agent"]}')
        
        # Set page load timeout
        chrome_options.page_load_strategy = 'normal'
        
        driver = None
        try:
            print("  Starting Chrome browser...")
            if use_manager:
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
            else:
                driver = webdriver.Chrome(options=chrome_options)
            
            # Set timeouts
            driver.set_page_load_timeout(30)  # 30 seconds for page load
            driver.set_script_timeout(30)      # 30 seconds for scripts
            
            print(f"  Loading {url}...")
            driver.get(url)
            
            if wait_for_table:
                print("  Waiting for table to load...")
                wait = WebDriverWait(driver, 15)
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            
            # Give extra time for dynamic content
            time.sleep(2)
            
            print("  Parsing page content...")
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            return soup
            
        except TimeoutException as e:
            raise Exception(f"Page load timeout: {str(e)}")
        except WebDriverException as e:
            raise Exception(f"WebDriver error: {str(e)}")
        except Exception as e:
            raise Exception(f"Selenium error: {str(e)}")
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
    
    def _clean_rate(self, rate_str):
        """Clean and convert rate string to float"""
        # Remove %, commas, and other characters
        cleaned = ''.join(c for c in rate_str if c.isdigit() or c in '.-')
        try:
            return float(cleaned)
        except ValueError:
            return None


class DataUpdater:
    """Handles updating Excel master files"""
    
    def __init__(self, input_dir="../swap_implied_input"):
        self.input_dir = Path(input_dir)
        self.files = {
            '1M': self.input_dir / 'input_master_1m.xlsx',
            '3M': self.input_dir / 'input_master_3m.xlsx',
            '6M': self.input_dir / 'input_master_6m.xlsx'
        }
    
    def validate_files(self):
        """Check if all master files exist"""
        missing_files = []
        for period, filepath in self.files.items():
            if not filepath.exists():
                missing_files.append(str(filepath))
        
        if missing_files:
            print(f"✗ Missing files: {', '.join(missing_files)}")
            return False
        
        print("✓ All master files found")
        return True
    
    def update_files(self, sofr_rates, fx_rate, forward_points):
        """
        Append new data to all master files
        
        Args:
            sofr_rates: dict with keys '1M', '3M', '6M'
            fx_rate: float
            forward_points: dict with keys '1M', '3M', '6M'
        """
        today = datetime.now().strftime('%Y-%m-%d')
        
        results = {}
        for period in ['1M', '3M', '6M']:
            filepath = self.files[period]
            
            try:
                print(f"\nUpdating {filepath.name}...")
                
                # Read existing data
                df = pd.read_excel(filepath)
                
                # Validate columns
                expected_cols = ['Date', f'{period[0]}mSOFR', 'USDSGD_FX', 'ForwardPoints']
                if list(df.columns) != expected_cols:
                    print(f"  Warning: Column mismatch. Expected {expected_cols}, got {list(df.columns)}")
                
                # Check if today's data already exists
                if 'Date' in df.columns:
                    df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
                    if today in df['Date'].values:
                        print(f"  ⚠ Data for {today} already exists. Skipping...")
                        results[period] = 'skipped'
                        continue
                
                # Create new row
                new_row = {
                    'Date': today,
                    f'{period[0]}mSOFR': sofr_rates[period],
                    'USDSGD_FX': fx_rate,
                    'ForwardPoints': forward_points[period]
                }
                
                # Append new row
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                
                # Save back to file
                df.to_excel(filepath, index=False)
                
                print(f"  ✓ Successfully appended: Date={today}, SOFR={sofr_rates[period]}%, FX={fx_rate}, FP={forward_points[period]}")
                results[period] = 'success'
                
            except Exception as e:
                print(f"  ✗ Error updating {filepath.name}: {e}")
                results[period] = 'failed'
        
        return results


def manual_forward_points_input():
    """Get forward points via manual input"""
    print("\n" + "=" * 70)
    print("MANUAL FORWARD POINTS INPUT")
    print("=" * 70)
    print("\nPlease enter BID and ASK for each tenor:")
    print("(Mid rate will be calculated automatically)\n")
    
    try:
        fp_1m_bid = float(input("1M Bid: "))
        fp_1m_ask = float(input("1M Ask: "))
        fp_3m_bid = float(input("3M Bid: "))
        fp_3m_ask = float(input("3M Ask: "))
        fp_6m_bid = float(input("6M Bid: "))
        fp_6m_ask = float(input("6M Ask: "))
        
        forward_points = {
            '1M': round((fp_1m_bid + fp_1m_ask) / 2, 4),
            '3M': round((fp_3m_bid + fp_3m_ask) / 2, 4),
            '6M': round((fp_6m_bid + fp_6m_ask) / 2, 4)
        }
        
        print(f"\n✓ Calculated mid rates: 1M={forward_points['1M']}, 3M={forward_points['3M']}, 6M={forward_points['6M']}")
        return forward_points
    except (ValueError, KeyboardInterrupt):
        print("\n✗ Invalid input or cancelled")
        return None


def calculate_forward_points(spot_rate, sofr_rates, sora_rates=None):
    """Calculate forward points from interest rates"""
    if sora_rates is None:
        print("⚠ SORA rates not provided, estimating as SOFR - 0.5%")
        sora_rates = {k: v - 0.5 for k, v in sofr_rates.items()}
    
    days_map = {'1M': 30, '3M': 90, '6M': 180}
    forward_points = {}
    
    for period, days in days_map.items():
        sofr = sofr_rates[period] / 100  # Convert to decimal
        sora = sora_rates[period] / 100
        forward_rate = spot_rate * ((1 + sofr * days / 360) / (1 + sora * days / 360))
        fp = (forward_rate - spot_rate) * 10000
        forward_points[period] = round(fp, 4)
    
    print(f"✓ Calculated forward points: 1M={forward_points['1M']}, 3M={forward_points['3M']}, 6M={forward_points['6M']}")
    return forward_points


def main():
    parser = argparse.ArgumentParser(
        description='Update swap-implied rate master files with latest data',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--selenium',
        action='store_true',
        help='Use Selenium for web scraping (more reliable but slower)'
    )
    parser.add_argument(
        '--input-dir',
        default='../swap_implied_input',
        help='Directory containing master Excel files (default: ../swap_implied_input)'
    )
    parser.add_argument(
        '--manual',
        action='store_true',
        help='Use manual input mode for forward points (when automated extraction fails)'
    )
    parser.add_argument(
        '--calculate',
        action='store_true',
        help='Calculate forward points from SOFR+SORA rates instead of scraping'
    )
    parser.add_argument(
        '--create-sample',
        action='store_true',
        help='Create sample master files for testing'
    )
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("SWAP-IMPLIED RATE DATA UPDATER")
    print("=" * 70)
    print()
    
    # Create sample files if requested
    if args.create_sample:
        create_sample_files(args.input_dir)
        return 0
    
    # Initialize
    extractor = DataExtractor(use_selenium=args.selenium)
    updater = DataUpdater(input_dir=args.input_dir)
    
    # Validate files exist
    if not updater.validate_files():
        print("\nRun with --create-sample to create sample master files")
        return 1
    
    print()
    print("EXTRACTING DATA FROM SOURCES")
    print("-" * 70)
    
    # Extract data from all sources
    sofr_rates = extractor.extract_sofr_rates()
    fx_rate = extractor.extract_usdsgd_fx()
    
    # Handle forward points based on mode
    if args.manual:
        print("\n" + "-" * 70)
        print("MANUAL INPUT MODE FOR FORWARD POINTS")
        print("-" * 70)
        forward_points = manual_forward_points_input()
    elif args.calculate:
        print("\n" + "-" * 70)
        print("CALCULATING FORWARD POINTS FROM INTEREST RATES")
        print("-" * 70)
        if sofr_rates and fx_rate:
            forward_points = calculate_forward_points(fx_rate, sofr_rates)
        else:
            print("✗ Cannot calculate: Missing SOFR rates or FX rate")
            forward_points = None
    else:
        forward_points = extractor.extract_forward_points()
        
        # If extraction failed, offer manual input
        if not forward_points:
            print("\n" + "=" * 70)
            print("⚠ AUTOMATED EXTRACTION FAILED")
            print("=" * 70)
            print("\nWould you like to:")
            print("  1. Enter forward points manually")
            print("  2. Calculate from SOFR rates (estimated)")
            print("  3. Skip forward points update")
            
            choice = input("\nChoice (1-3): ").strip()
            
            if choice == '1':
                forward_points = manual_forward_points_input()
            elif choice == '2' and sofr_rates and fx_rate:
                forward_points = calculate_forward_points(fx_rate, sofr_rates)
            else:
                print("Skipping forward points...")

    
    # Check if all extractions succeeded
    if not all([sofr_rates, fx_rate, forward_points]):
        print("\n✗ Failed to extract all required data")
        if not args.selenium:
            print("\nTip: Try running with --selenium flag for more reliable extraction")
        return 1
    
    print()
    print("UPDATING MASTER FILES")
    print("-" * 70)
    
    # Update files
    results = updater.update_files(sofr_rates, fx_rate, forward_points)
    
    # Summary
    print()
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    success_count = sum(1 for r in results.values() if r == 'success')
    print(f"Successfully updated: {success_count}/3 files")
    for period, status in results.items():
        symbol = "✓" if status == "success" else "⚠" if status == "skipped" else "✗"
        print(f"  {symbol} {period}: {status}")
    print("=" * 70)
    
    return 0 if success_count > 0 else 1


def create_sample_files(input_dir="../swap_implied_input"):
    """Create sample master files for testing"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    Path(input_dir).mkdir(parents=True, exist_ok=True)
    
    print(f"Creating sample master files in {input_dir}...")
    
    for period in ['1M', '3M', '6M']:
        filename = f'input_master_{period.lower()}.xlsx'
        filepath = Path(input_dir) / filename
        
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Data'
        
        # Headers
        headers = ['Date', f'{period[0]}mSOFR', 'USDSGD_FX', 'ForwardPoints']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['2025-02-01', 4.50, 1.3450, -25.50],
            ['2025-02-02', 4.51, 1.3455, -25.60],
            ['2025-02-03', 4.49, 1.3448, -25.45],
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        # Column widths
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 15
        
        wb.save(filepath)
        print(f"  ✓ Created {filename}")
    
    print(f"\nSample files created successfully in {input_dir}/")
    print("You can now run the script normally to update these files.")


if __name__ == "__main__":
    sys.exit(main())
